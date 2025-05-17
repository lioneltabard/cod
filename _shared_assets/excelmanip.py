import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

def choisir_fichier():
    """Ouvre une boîte de dialogue pour choisir un fichier Excel."""
    root = tk.Tk()
    root.withdraw()  # Cacher la fenêtre principale
    chemin_fichier = filedialog.askopenfilename(
        title="Sélectionnez un fichier Excel",
        filetypes=[("Fichiers Excel", "*.xlsx *.xls")]
    )
    root.destroy()
    return chemin_fichier

def main():
    # Étape 1: Choisir un fichier Excel
    print("Veuillez sélectionner un fichier Excel...")
    chemin_fichier = choisir_fichier()
    if not chemin_fichier:
        print("Aucun fichier sélectionné. Fin du programme.")
        return
    
    # Charger le fichier avec openpyxl pour pouvoir le modifier
    classeur = openpyxl.load_workbook(chemin_fichier)
    
    # Étape 2: Choisir l'onglet
    print("\nOnglets disponibles :")
    for i, nom_onglet in enumerate(classeur.sheetnames, 1):
        print(f"{i}. {nom_onglet}")
    
    choix_onglet = 0
    while choix_onglet < 1 or choix_onglet > len(classeur.sheetnames):
        try:
            choix_onglet = int(input("\nEntrez le numéro de l'onglet que vous souhaitez utiliser : "))
        except ValueError:
            print("Veuillez entrer un nombre valide.")
    
    nom_onglet = classeur.sheetnames[choix_onglet - 1]
    feuille = classeur[nom_onglet]
    
    # Étape 3: Choisir la colonne pour rechercher les cellules vides
    print("\nColonnes disponibles :")
    colonnes_info = []
    for i in range(1, feuille.max_column + 1):
        cellule = feuille.cell(row=1, column=i)
        col_letter = get_column_letter(i)
        colonnes_info.append(f"{col_letter}. {cellule.value}")
    
    # Afficher toutes les colonnes sur une seule ligne
    print(" | ".join(colonnes_info))
    
    colonne_recherche = input("\nEntrez la lettre de la colonne dans laquelle vous souhaitez rechercher des cellules vides : ").upper()
    col_index = column_index_from_string(colonne_recherche)
    
    # Étape 4-7: Itérer à travers les lignes, trouver les cellules vides,
    # demander une saisie, écrire la valeur et sauvegarder
    continuer = True
    for row in range(2, feuille.max_row + 1):  # Commencer à la ligne 2 (après les en-têtes)
        # Vérifier si la cellule est vide
        cellule = feuille.cell(row=row, column=col_index)
        if cellule.value is None or cellule.value == "":
            print(f"\nLigne {row} avec cellule vide dans la colonne {colonne_recherche}:")
            
            # Afficher les autres colonnes de cette ligne (sur une seule ligne)
            infos_ligne = []
            for col in range(1, feuille.max_column + 1):
                if col != col_index:
                    valeur = feuille.cell(row=row, column=col).value
                    en_tete = feuille.cell(row=1, column=col).value
                    infos_ligne.append(f"{en_tete}: {valeur}")
            
            # Afficher toutes les informations sur une seule ligne
            print(" | ".join(infos_ligne))
            
            # Demander la saisie
            nouvelle_valeur = input(f"\nEntrez la valeur pour la cellule vide ({colonne_recherche}{row}): ")
            
            # Écrire la valeur dans la cellule
            feuille.cell(row=row, column=col_index).value = nouvelle_valeur
            
            # Sauvegarder le fichier
            classeur.save(chemin_fichier)
            print("Fichier sauvegardé!")
            
            # Demander si l'utilisateur veut continuer ou quitter
            choix = input("\nVoulez-vous continuer? (o/n): ").lower()
            if choix != 'o' and choix != 'oui':
                continuer = False
                break
    
    if continuer:
        print("\nToutes les cellules vides dans la colonne sélectionnée ont été traitées.")
    
    print("Fin du programme.")

if __name__ == "__main__":
    main()
